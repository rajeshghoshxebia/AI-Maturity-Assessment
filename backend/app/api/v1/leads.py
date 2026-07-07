from __future__ import annotations

import io
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from pydantic import BaseModel, Field

from app.core.auth import CurrentUser, get_current_user
from app.services.lead_search import (
    ICPUnavailable,
    Lead,
    ProviderNotConfigured,
    SearchCriteria,
    derive_criteria,
    get_provider,
)

router = APIRouter()

# Guard rails: a search should never silently pull thousands of records.
_DEFAULT_LIMIT = 25
_MAX_LIMIT = 100


class LeadSearchRequest(BaseModel):
    business_case: str = Field(min_length=3, max_length=4000)
    company_name: str | None = None
    limit: int = _DEFAULT_LIMIT
    # When supplied, these override the LLM-derived criteria (lets the user edit
    # the ICP and re-run without re-deriving).
    criteria: SearchCriteria | None = None


class LeadSearchResponse(BaseModel):
    criteria: SearchCriteria
    leads: list[Lead]
    provider: str
    count: int


class LeadExportRequest(BaseModel):
    leads: list[Lead]
    business_case: str | None = None
    company_name: str | None = None


@router.post("/search", response_model=LeadSearchResponse)
async def search_leads(
    body: LeadSearchRequest,
    user: CurrentUser = Depends(get_current_user),
) -> LeadSearchResponse:
    limit = max(1, min(body.limit, _MAX_LIMIT))

    if body.criteria is not None:
        criteria = body.criteria
    else:
        try:
            criteria = await derive_criteria(body.business_case, body.company_name)
        except ICPUnavailable as exc:
            raise HTTPException(status_code=503, detail=str(exc))
        except Exception as exc:  # noqa: BLE001 - surface LLM failures cleanly
            raise HTTPException(
                status_code=502, detail=f"Failed to derive criteria: {type(exc).__name__}: {exc}"
            )

    try:
        provider = get_provider()
    except ProviderNotConfigured as exc:
        raise HTTPException(status_code=503, detail=str(exc))

    try:
        leads = await provider.search(criteria, limit=limit)
    except ProviderNotConfigured as exc:
        raise HTTPException(status_code=503, detail=str(exc))
    except Exception as exc:  # noqa: BLE001 - surface provider/API errors cleanly
        raise HTTPException(
            status_code=502, detail=f"Lead search failed: {type(exc).__name__}: {exc}"
        )

    return LeadSearchResponse(
        criteria=criteria,
        leads=leads,
        provider=provider.name,
        count=len(leads),
    )


_EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("full_name", "Name"),
    ("title", "Title"),
    ("seniority", "Seniority"),
    ("company", "Company"),
    ("industry", "Industry"),
    ("location", "Location"),
    ("email", "Email"),
    ("email_verified", "Email verified"),
    ("phone", "Phone"),
    ("linkedin_url", "LinkedIn"),
    ("source", "Source"),
]


@router.post("/export")
async def export_leads(
    body: LeadExportRequest,
    user: CurrentUser = Depends(get_current_user),
) -> Response:
    if not body.leads:
        raise HTTPException(status_code=400, detail="No leads to export.")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=503,
            detail="openpyxl is not installed. Run: pip install openpyxl",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F1149")  # Xebia blue-dark

    for col_idx, (_, header) in enumerate(_EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, lead in enumerate(body.leads, start=2):
        for col_idx, (attr, _) in enumerate(_EXPORT_COLUMNS, start=1):
            value = getattr(lead, attr)
            if isinstance(value, bool):
                value = "Yes" if value else "No"
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Reasonable column widths from the header + longest cell.
    for col_idx, (attr, header) in enumerate(_EXPORT_COLUMNS, start=1):
        longest = max(
            [len(header)] + [len(str(getattr(l, attr) or "")) for l in body.leads]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(longest + 2, 50)
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M")
    filename = f"leads-{stamp}.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
